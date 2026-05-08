"""
Students router — mounted at /api/students
Every query filters by school_id == current_user.current_school_id.
"""
from __future__ import annotations

import io
import logging
import uuid

import openpyxl  # type: ignore
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload

from app.database import get_db
from app.dependencies.rbac import require_any_permission, require_permission
from app.models.class_ import Class
from app.models.student import Gender, Student
from app.models.user import User
from app.schemas.student import (
    AssignClassRequest,
    BulkUploadResult,
    PaginatedStudentResponse,
    StudentCreate,
    StudentListItem,
    StudentPromotion,
    StudentResponse,
    StudentTransfer,
    StudentUpdate,
)
from app.services.student_service import generate_admission_number, process_bulk_upload

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/students", tags=["students"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _full_name(student: Student) -> str:
    parts = [student.first_name]
    if student.middle_name:
        parts.append(student.middle_name)
    parts.append(student.last_name)
    return " ".join(parts)


def _to_list_item(student: Student) -> StudentListItem:
    class_name: str | None = None
    if student.current_class is not None:
        class_name = student.current_class.name

    return StudentListItem(
        id=student.id,
        first_name=student.first_name,
        last_name=student.last_name,
        full_name=_full_name(student),
        admission_number=student.admission_number,
        class_name=class_name,
        gender=student.gender.value,
        is_active=student.is_active,
        admission_date=student.admission_date,
        photo_url=student.photo_url,
    )


def _to_response(student: Student) -> StudentResponse:
    class_name: str | None = None
    if student.current_class is not None:
        class_name = student.current_class.name

    parent_count = len(student.parents) if student.parents is not None else 0

    return StudentResponse(
        id=student.id,
        school_id=student.school_id,
        admission_number=student.admission_number,
        first_name=student.first_name,
        last_name=student.last_name,
        middle_name=student.middle_name,
        full_name=_full_name(student),
        date_of_birth=student.date_of_birth,
        gender=student.gender.value,
        address=student.address,
        state_of_origin=student.state_of_origin,
        religion=student.religion,
        blood_group=student.blood_group,
        genotype=student.genotype,
        class_id=student.class_id,
        class_name=class_name,
        is_active=student.is_active,
        admission_date=student.admission_date,
        photo_url=student.photo_url,
        parent_count=parent_count,
        created_at=student.created_at,
        updated_at=student.updated_at,
    )


async def _get_student_or_404(
    student_id: uuid.UUID,
    school_id: uuid.UUID,
    db: AsyncSession,
) -> Student:
    result = await db.execute(
        select(Student)
        .where(
            Student.id == student_id,
            Student.school_id == school_id,
        )
        .options(
            joinedload(Student.current_class),
            selectinload(Student.parents),
        )
    )
    student = result.scalar_one_or_none()
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    return student


# ---------------------------------------------------------------------------
# POST /students/ — create student
# ---------------------------------------------------------------------------

@router.post("/", response_model=StudentResponse, status_code=status.HTTP_201_CREATED)
async def create_student(
    body: StudentCreate,
    current_user: User = Depends(require_permission("create_student")),
    db: AsyncSession = Depends(get_db),
) -> StudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]

    # Validate class belongs to same school
    if body.class_id is not None:
        cls_result = await db.execute(
            select(Class).where(
                Class.id == body.class_id,
                Class.school_id == school_id,
            )
        )
        if cls_result.scalar_one_or_none() is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Class not found in this school",
            )

    # Resolve admission number
    if body.admission_number:
        # Check uniqueness within school
        dup = await db.execute(
            select(Student).where(
                Student.school_id == school_id,
                Student.admission_number == body.admission_number,
            )
        )
        if dup.scalar_one_or_none() is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Admission number '{body.admission_number}' already exists in this school",
            )
        adm_num = body.admission_number
    else:
        adm_num = await generate_admission_number(db, school_id)

    student = Student(
        school_id=school_id,
        admission_number=adm_num,
        first_name=body.first_name,
        last_name=body.last_name,
        middle_name=body.middle_name,
        date_of_birth=body.date_of_birth,
        gender=Gender(body.gender),
        address=body.address,
        state_of_origin=body.state_of_origin,
        religion=body.religion,
        blood_group=body.blood_group,
        genotype=body.genotype,
        class_id=body.class_id,
        admission_date=body.admission_date,
        photo_url=body.photo_url,
        is_active=True,
    )
    db.add(student)
    await db.commit()
    await db.refresh(student)

    # Re-fetch with relationships
    student = await _get_student_or_404(student.id, school_id, db)
    return _to_response(student)


# ---------------------------------------------------------------------------
# GET /students/ — list students (admin/super_admin)
# ---------------------------------------------------------------------------

@router.get("/", response_model=PaginatedStudentResponse)
async def list_students(
    search: str | None = Query(None),
    class_id: uuid.UUID | None = Query(None),
    gender: str | None = Query(None),
    is_active: bool | None = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: User = Depends(require_permission("view_all_students")),
    db: AsyncSession = Depends(get_db),
) -> PaginatedStudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]

    filters = [Student.school_id == school_id]

    if search:
        term = f"%{search}%"
        filters.append(
            or_(
                Student.first_name.ilike(term),
                Student.last_name.ilike(term),
                Student.middle_name.ilike(term),
                Student.admission_number.ilike(term),
            )
        )
    if class_id is not None:
        filters.append(Student.class_id == class_id)
    if gender is not None:
        filters.append(Student.gender == Gender(gender))
    if is_active is not None:
        filters.append(Student.is_active == is_active)

    total_result = await db.execute(
        select(func.count(Student.id)).where(and_(*filters))
    )
    total: int = total_result.scalar_one()

    result = await db.execute(
        select(Student)
        .where(and_(*filters))
        .options(joinedload(Student.current_class))
        .order_by(Student.last_name, Student.first_name)
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    students = result.scalars().all()

    return PaginatedStudentResponse(
        items=[_to_list_item(s) for s in students],
        total=total,
        page=page,
        per_page=per_page,
        total_pages=max(1, (total + per_page - 1) // per_page),
    )


# ---------------------------------------------------------------------------
# GET /students/my-class — teacher view of own class
# ---------------------------------------------------------------------------

@router.get("/my-class", response_model=PaginatedStudentResponse)
async def list_my_class_students(
    search: str | None = Query(None),
    is_active: bool | None = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: User = Depends(require_permission("view_own_class_students")),
    db: AsyncSession = Depends(get_db),
) -> PaginatedStudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]
    teacher_class_id: uuid.UUID | None = current_user.current_class_id  # type: ignore[attr-defined]

    if teacher_class_id is None:
        return PaginatedStudentResponse(items=[], total=0, page=page, per_page=per_page, total_pages=1)

    # Verify the class belongs to this school
    cls_result = await db.execute(
        select(Class).where(
            Class.id == teacher_class_id,
            Class.school_id == school_id,
        )
    )
    if cls_result.scalar_one_or_none() is None:
        return PaginatedStudentResponse(items=[], total=0, page=page, per_page=per_page, total_pages=1)

    filters = [
        Student.school_id == school_id,
        Student.class_id == teacher_class_id,
    ]
    if search:
        term = f"%{search}%"
        filters.append(
            or_(
                Student.first_name.ilike(term),
                Student.last_name.ilike(term),
                Student.admission_number.ilike(term),
            )
        )
    if is_active is not None:
        filters.append(Student.is_active == is_active)

    total_result = await db.execute(
        select(func.count(Student.id)).where(and_(*filters))
    )
    total: int = total_result.scalar_one()

    result = await db.execute(
        select(Student)
        .where(and_(*filters))
        .options(joinedload(Student.current_class))
        .order_by(Student.last_name, Student.first_name)
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    students = result.scalars().all()

    return PaginatedStudentResponse(
        items=[_to_list_item(s) for s in students],
        total=total,
        page=page,
        per_page=per_page,
        total_pages=max(1, (total + per_page - 1) // per_page),
    )


# ---------------------------------------------------------------------------
# GET /students/{student_id}
# ---------------------------------------------------------------------------

@router.get("/{student_id}", response_model=StudentResponse)
async def get_student(
    student_id: uuid.UUID,
    current_user: User = Depends(require_any_permission("view_all_students", "view_own_class_students")),
    db: AsyncSession = Depends(get_db),
) -> StudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]
    student = await _get_student_or_404(student_id, school_id, db)

    # If teacher role, restrict to own class only
    role = current_user.current_role.value  # type: ignore[attr-defined]
    if role == "teacher":
        teacher_class_id = current_user.current_class_id  # type: ignore[attr-defined]
        if student.class_id != teacher_class_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only view students in your own class",
            )

    return _to_response(student)


# ---------------------------------------------------------------------------
# PATCH /students/{student_id}
# ---------------------------------------------------------------------------

@router.patch("/{student_id}", response_model=StudentResponse)
async def update_student(
    student_id: uuid.UUID,
    body: StudentUpdate,
    current_user: User = Depends(require_permission("edit_student")),
    db: AsyncSession = Depends(get_db),
) -> StudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]
    student = await _get_student_or_404(student_id, school_id, db)

    if body.class_id is not None:
        cls_result = await db.execute(
            select(Class).where(
                Class.id == body.class_id,
                Class.school_id == school_id,
            )
        )
        if cls_result.scalar_one_or_none() is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Class not found in this school",
            )

    update_data = body.model_dump(exclude_unset=True)
    if "gender" in update_data:
        update_data["gender"] = Gender(update_data["gender"])

    for field, value in update_data.items():
        setattr(student, field, value)

    await db.commit()
    await db.refresh(student)
    student = await _get_student_or_404(student.id, school_id, db)
    return _to_response(student)


# ---------------------------------------------------------------------------
# DELETE /students/{student_id}/deactivate
# ---------------------------------------------------------------------------

@router.delete("/{student_id}/deactivate", response_model=StudentResponse)
async def deactivate_student(
    student_id: uuid.UUID,
    current_user: User = Depends(require_permission("edit_student")),
    db: AsyncSession = Depends(get_db),
) -> StudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]
    student = await _get_student_or_404(student_id, school_id, db)
    student.is_active = False
    await db.commit()
    await db.refresh(student)
    student = await _get_student_or_404(student.id, school_id, db)
    return _to_response(student)


# ---------------------------------------------------------------------------
# POST /students/{student_id}/assign-class
# ---------------------------------------------------------------------------

@router.post("/{student_id}/assign-class", response_model=StudentResponse)
async def assign_class(
    student_id: uuid.UUID,
    body: AssignClassRequest,
    current_user: User = Depends(require_permission("assign_student_to_class")),
    db: AsyncSession = Depends(get_db),
) -> StudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]

    cls_result = await db.execute(
        select(Class).where(Class.id == body.class_id, Class.school_id == school_id)
    )
    if cls_result.scalar_one_or_none() is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Class not found in this school")

    student = await _get_student_or_404(student_id, school_id, db)
    student.class_id = body.class_id
    await db.commit()
    student = await _get_student_or_404(student.id, school_id, db)
    return _to_response(student)


# ---------------------------------------------------------------------------
# POST /students/{student_id}/remove-class
# ---------------------------------------------------------------------------

@router.post("/{student_id}/remove-class", response_model=StudentResponse)
async def remove_class(
    student_id: uuid.UUID,
    current_user: User = Depends(require_permission("remove_student_from_class")),
    db: AsyncSession = Depends(get_db),
) -> StudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]
    student = await _get_student_or_404(student_id, school_id, db)
    student.class_id = None
    await db.commit()
    student = await _get_student_or_404(student.id, school_id, db)
    return _to_response(student)


# ---------------------------------------------------------------------------
# POST /students/{student_id}/transfer
# ---------------------------------------------------------------------------

@router.post("/{student_id}/transfer", response_model=StudentResponse)
async def transfer_student(
    student_id: uuid.UUID,
    body: StudentTransfer,
    current_user: User = Depends(require_permission("transfer_student")),
    db: AsyncSession = Depends(get_db),
) -> StudentResponse:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]

    cls_result = await db.execute(
        select(Class).where(Class.id == body.new_class_id, Class.school_id == school_id)
    )
    if cls_result.scalar_one_or_none() is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Target class not found in this school")

    student = await _get_student_or_404(student_id, school_id, db)
    student.class_id = body.new_class_id
    await db.commit()
    student = await _get_student_or_404(student.id, school_id, db)
    return _to_response(student)


# ---------------------------------------------------------------------------
# POST /students/promote — bulk promote/repeat
# ---------------------------------------------------------------------------

@router.post("/promote", status_code=status.HTTP_200_OK)
async def promote_students(
    body: StudentPromotion,
    current_user: User = Depends(require_permission("promote_student")),
    db: AsyncSession = Depends(get_db),
) -> dict:
    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]

    if body.action.value == "promote" and body.new_class_id is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="new_class_id is required when action is 'promote'",
        )

    if body.action.value == "promote":
        cls_result = await db.execute(
            select(Class).where(Class.id == body.new_class_id, Class.school_id == school_id)
        )
        if cls_result.scalar_one_or_none() is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Target class not found in this school")

    # Verify all students belong to this school in one query
    owned_result = await db.execute(
        select(func.count(Student.id)).where(
            Student.id.in_(body.student_ids),
            Student.school_id == school_id,
        )
    )
    owned_count = owned_result.scalar_one()
    if owned_count != len(body.student_ids):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="One or more students do not belong to this school",
        )

    if body.action.value == "promote":
        await db.execute(
            update(Student)
            .where(Student.id.in_(body.student_ids), Student.school_id == school_id)
            .values(class_id=body.new_class_id)
        )
    else:
        # repeat — remove from class
        await db.execute(
            update(Student)
            .where(Student.id.in_(body.student_ids), Student.school_id == school_id)
            .values(class_id=None)
        )

    await db.commit()
    return {"updated_count": len(body.student_ids), "action": body.action.value}


# ---------------------------------------------------------------------------
# POST /students/bulk-upload
# ---------------------------------------------------------------------------

@router.post("/bulk-upload", response_model=BulkUploadResult)
async def bulk_upload_students(
    file: UploadFile = File(...),
    current_user: User = Depends(require_permission("bulk_upload_students")),
    db: AsyncSession = Depends(get_db),
) -> BulkUploadResult:
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are supported",
        )

    school_id: uuid.UUID = current_user.current_school_id  # type: ignore[attr-defined]
    return await process_bulk_upload(db, file, school_id, current_user)


# ---------------------------------------------------------------------------
# GET /students/bulk-upload/template
# ---------------------------------------------------------------------------

@router.get("/bulk-upload/template")
async def bulk_upload_template(
    current_user: User = Depends(require_permission("bulk_upload_students")),
) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "First Name*", "Last Name*", "Middle Name",
        "Date of Birth* (YYYY-MM-DD)", "Gender* (male/female)",
        "Admission Number", "Admission Date* (YYYY-MM-DD)",
        "Class Name", "Address", "State of Origin",
    ]
    ws.append(headers)

    # Example rows
    ws.append([
        "Chukwuemeka", "Okafor", "Ifeanyi",
        "2012-03-15", "male",
        "", "2023-09-01",
        "JSS 1A", "12 Lagos Road, Ikeja", "Anambra",
    ])
    ws.append([
        "Amina", "Bello", "",
        "2011-07-22", "female",
        "SCH-2023-0002", "2023-09-01",
        "JSS 2B", "", "Kano",
    ])

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.column_letter  # trigger dimension tracking

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_template.xlsx"},
    )
